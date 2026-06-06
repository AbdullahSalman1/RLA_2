from __future__ import annotations

from pathlib import Path

import requests
from openpyxl import load_workbook

from geocache import geocode_all as geocode_all_cached

# ── API Key ───────────────────────────────────────────────────────────────────
ORS_API_KEY = "eyJvcmciOiI1YjNjZTM1OTc4NTExMTAwMDFjZjYyNDgiLCJpZCI6Ijc4MzMyNTBkYTBhMDRiYTg5MDQyYzkxNTQ4MDY0MzQ4IiwiaCI6Im11cm11cjY0In0="

DATA_FILE = Path(__file__).with_name("delivery_data.xlsx")
FLEET_FILE = Path(__file__).with_name("fleet_data.xlsx")

WAREHOUSE_DEFAULT = {"id": "W0", "name": "Warehouse", "address": "14 Rue du Chemin Vert, Paris"}
VEHICLES_DEFAULT = [
    {"id": "V1", "type": "small", "label": "Small Van 1", "max_stops": 30},
    {"id": "V2", "type": "small", "label": "Small Van 2", "max_stops": 30},
    {"id": "V3", "type": "small", "label": "Small Van 3", "max_stops": 30},
    {"id": "V4", "type": "large", "label": "Large Van 1", "max_stops": 20},
    {"id": "V5", "type": "large", "label": "Large Van 2", "max_stops": 20},
    {"id": "V6", "type": "refrigerated", "label": "Refrigerated Van", "max_stops": 15},
]
DRIVERS_DEFAULT = [
    {"id": "D1", "name": "Alexandre M.", "certified_refrigerated": True, "shift_start": "06:30", "max_hours": 8},
    {"id": "D2", "name": "Fatima B.", "certified_refrigerated": True, "shift_start": "06:30", "max_hours": 8},
    {"id": "D3", "name": "Thomas L.", "certified_refrigerated": False, "shift_start": "07:00", "max_hours": 8},
    {"id": "D4", "name": "Yasmine K.", "certified_refrigerated": False, "shift_start": "07:00", "max_hours": 8},
    {"id": "D5", "name": "Romain D.", "certified_refrigerated": False, "shift_start": "07:00", "max_hours": 8},
]

SERVICE_TIME_MIN = 10


def _as_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "oui", "x"}


def load_delivery_data(path=DATA_FILE):
    """Load warehouse and orders from delivery_data.xlsx."""
    if not path.exists():
        return WAREHOUSE_DEFAULT.copy(), []

    wb = load_workbook(path, data_only=True)
    if "warehouse" not in wb.sheetnames or "orders" not in wb.sheetnames:
        return WAREHOUSE_DEFAULT.copy(), []

    ws_wh = wb["warehouse"]
    wh_rows = list(ws_wh.iter_rows(values_only=True))
    wh_headers = [str(v).strip().lower() for v in wh_rows[0]]
    wh_data = wh_rows[1]
    wh_map = {name: idx for idx, name in enumerate(wh_headers)}

    warehouse = {
        "id": str(wh_data[wh_map["id"]]),
        "name": str(wh_data[wh_map["name"]]),
        "address": str(wh_data[wh_map["address"]]),
    }

    ws_orders = wb["orders"]
    order_rows = list(ws_orders.iter_rows(values_only=True))
    order_headers = [str(v).strip().lower() for v in order_rows[0]]
    order_map = {name: idx for idx, name in enumerate(order_headers)}

    orders = []
    for row in order_rows[1:]:
        if not any(row):
            continue
        orders.append({
            "id": str(row[order_map["id"]]),
            "name": str(row[order_map["name"]]),
            "address": str(row[order_map["address"]]),
            "priority": str(row[order_map["priority"]]).lower(),
            "is_cold": _as_bool(row[order_map["is_cold"]]),
            "is_suburban": _as_bool(row[order_map["is_suburban"]]),
            "boxes": int(row[order_map["boxes"]]),
            "time_window": (
                str(row[order_map["time_window_start"]]),
                str(row[order_map["time_window_end"]]),
            ),
        })

    return warehouse, orders


def load_fleet_data(path=FLEET_FILE):
    """Load vehicles and drivers from fleet_data.xlsx."""
    if not path.exists():
        return VEHICLES_DEFAULT.copy(), DRIVERS_DEFAULT.copy()

    wb = load_workbook(path, data_only=True)
    if "vehicles" not in wb.sheetnames or "drivers" not in wb.sheetnames:
        return VEHICLES_DEFAULT.copy(), DRIVERS_DEFAULT.copy()

    ws_vehicles = wb["vehicles"]
    vehicle_rows = list(ws_vehicles.iter_rows(values_only=True))
    vehicle_headers = [str(v).strip().lower() for v in vehicle_rows[0]]
    vehicle_map = {name: idx for idx, name in enumerate(vehicle_headers)}

    vehicles = []
    for row in vehicle_rows[1:]:
        if not any(row):
            continue
        vehicles.append({
            "id": str(row[vehicle_map["id"]]),
            "type": str(row[vehicle_map["type"]]).lower(),
            "label": str(row[vehicle_map["label"]]),
            "max_stops": int(row[vehicle_map["max_stops"]]),
        })

    ws_drivers = wb["drivers"]
    driver_rows = list(ws_drivers.iter_rows(values_only=True))
    driver_headers = [str(v).strip().lower() for v in driver_rows[0]]
    driver_map = {name: idx for idx, name in enumerate(driver_headers)}

    drivers = []
    for row in driver_rows[1:]:
        if not any(row):
            continue
        drivers.append({
            "id": str(row[driver_map["id"]]),
            "name": str(row[driver_map["name"]]),
            "certified_refrigerated": _as_bool(row[driver_map["certified_refrigerated"]]),
            "shift_start": str(row[driver_map["shift_start"]]),
            "max_hours": int(row[driver_map["max_hours"]]),
        })

    return vehicles, drivers


WAREHOUSE, ORDERS = load_delivery_data()
VEHICLES, DRIVERS = load_fleet_data()


# ── Step 1 — Geocoding ───────────────────────────────────────────────────────
def geocode(name, address):
    """Convert address to lat/lon using ORS Geocoding API."""
    url = "https://api.openrouteservice.org/geocode/search"
    headers = {"Authorization": ORS_API_KEY}
    params = {"text": address, "boundary.country": "FR", "size": 1}

    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200 or not response.json().get("features"):
        print(f"  Could not geocode: {address}")
        return None

    feature = response.json()["features"][0]
    lon, lat = feature["geometry"]["coordinates"]
    return {"name": name, "address": address, "lat": lat, "lon": lon}


def geocode_all(orders):
    """Geocode warehouse and all orders."""
    print("\n" + "=" * 65)
    print("  STEP 1 — GEOCODING ADDRESSES")
    print("=" * 65)

    wh = geocode(WAREHOUSE["name"], WAREHOUSE["address"])
    wh["id"] = "W0"
    wh["priority"] = None
    wh["time_window"] = None
    wh["boxes"] = 0
    wh["is_cold"] = False
    wh["is_suburban"] = False
    print(f"  ✓ Warehouse: ({wh['lat']:.4f}, {wh['lon']:.4f})")

    locations = [wh]
    for order in orders:
        loc = geocode(order["name"], order["address"])
        if loc:
            loc["id"] = order["id"]
            loc["priority"] = order["priority"]
            loc["time_window"] = order["time_window"]
            loc["boxes"] = order["boxes"]
            loc["is_cold"] = order["is_cold"]
            loc["is_suburban"] = order["is_suburban"]
            locations.append(loc)
            print(f"  ✓ {order['name']}: ({loc['lat']:.4f}, {loc['lon']:.4f})")

    return locations


# ── Step 2 — Distance matrix ─────────────────────────────────────────────────
def get_distance_matrix(locations):
    print("\n" + "=" * 65)
    print("  STEP 2 — FETCHING DISTANCE MATRIX")
    print("=" * 65)

    url = "https://api.openrouteservice.org/v2/matrix/driving-car"
    headers = {"Authorization": ORS_API_KEY, "Content-Type": "application/json"}
    coords = [[loc["lon"], loc["lat"]] for loc in locations]
    body = {"locations": coords, "metrics": ["distance", "duration"], "units": "km"}

    print(f"  Calling ORS Matrix API with {len(locations)} locations...")
    response = requests.post(url, headers=headers, json=body)
    if response.status_code != 200:
        print(f"  API Error {response.status_code}: {response.text}")
        return None, None

    data = response.json()
    distances = data["distances"]
    durations = [[s / 60 for s in row] for row in data["durations"]]
    print(f"  ✓ Matrix ready: {len(locations)}x{len(locations)} pairs")
    return distances, durations


# ── Step 3 — Assignment ───────────────────────────────────────────────────────
def assign_orders_to_vehicles(orders):
    priority_rank = {"critical": 0, "high": 1, "normal": 2}
    sorted_orders = sorted(orders, key=lambda o: priority_rank.get(o["priority"], 99))
    vehicle_stops = {v["id"]: 0 for v in VEHICLES}
    vehicle_orders = {v["id"]: [] for v in VEHICLES}

    print("\n" + "=" * 65)
    print("  STEP 3 — ASSIGNING ORDERS TO VEHICLES")
    print("=" * 65)

    for order in sorted_orders:
        vid = None
        reason = ""

        if order["is_cold"]:
            preferred = ["refrigerated"]
            reason = "cold chain → refrigerated"
        elif order["is_suburban"]:
            preferred = ["large", "small", "refrigerated"]
            reason = "suburban → large van"
        else:
            preferred = ["small", "large", "refrigerated"]
            reason = "city → small van"

        for vehicle in VEHICLES:
            if vehicle["type"] in preferred and vehicle_stops[vehicle["id"]] < vehicle["max_stops"]:
                vid = vehicle["id"]
                break
        if vid is None:
            for vehicle in VEHICLES:
                if vehicle_stops[vehicle["id"]] < vehicle["max_stops"]:
                    vid = vehicle["id"]
                    reason = "assigned by remaining capacity"
                    break

        if vid:
            vehicle_orders[vid].append(order)
            vehicle_stops[vid] += 1
            vlabel = next(v["label"] for v in VEHICLES if v["id"] == vid)
            print(f"  {order['id']} {order['name']:<28} → {vlabel} ({reason})")
        else:
            print(f"  {order['id']} {order['name']:<28} → ⚠ UNASSIGNED")

    return vehicle_orders


# ── Step 4 — Driver assignment ───────────────────────────────────────────────
def assign_drivers(vehicle_orders):
    assigned_drivers = set()
    driver_assignment = {}

    for vehicle in VEHICLES:
        vid = vehicle["id"]
        if not vehicle_orders.get(vid):
            continue

        driver = None
        if vehicle["type"] == "refrigerated":
            for d in DRIVERS:
                if d["certified_refrigerated"] and d["id"] not in assigned_drivers:
                    driver = d
                    assigned_drivers.add(d["id"])
                    break

        if driver is None:
            for d in DRIVERS:
                if d["id"] not in assigned_drivers:
                    driver = d
                    assigned_drivers.add(d["id"])
                    break

        driver_assignment[vid] = driver

    return driver_assignment


# ── Routing helpers ──────────────────────────────────────────────────────────
def time_to_min(t):
    h, m = map(int, t.split(":"))
    return h * 60 + m


def min_to_time(m):
    return f"{int(m) // 60:02d}:{int(m) % 60:02d}"


def _nearest_neighbor_route(stop_indices, distances, start_idx=0):
    if not stop_indices:
        return []

    remaining = list(stop_indices)
    route = []
    current_pos = start_idx

    while remaining:
        chosen = min(remaining, key=lambda i: distances[current_pos][i])
        route.append(chosen)
        remaining.remove(chosen)
        current_pos = chosen

    return route


def build_route(stop_indices, locations, distances, driver=None, durations=None):
    """Critical stops first, then nearest-neighbor routing for the rest."""
    if not stop_indices:
        return [], 0

    critical = [i for i in stop_indices if locations[i]["priority"] == "critical"]
    others = [i for i in stop_indices if locations[i]["priority"] != "critical"]

    critical_route = _nearest_neighbor_route(critical, distances, start_idx=0)
    current_pos = critical_route[-1] if critical_route else 0
    other_route = _nearest_neighbor_route(others, distances, start_idx=current_pos)

    final_route = critical_route + other_route
    return final_route, len(critical_route)


def build_route_without_optimization(stop_indices, locations):
    critical = [i for i in stop_indices if locations[i]["priority"] == "critical"]
    high = [i for i in stop_indices if locations[i]["priority"] == "high"]
    normal = [i for i in stop_indices if locations[i]["priority"] == "normal"]

    critical.sort(key=lambda i: (time_to_min(locations[i]["time_window"][0]), time_to_min(locations[i]["time_window"][1])))
    high.sort(key=lambda i: (time_to_min(locations[i]["time_window"][0]), time_to_min(locations[i]["time_window"][1])))

    final_route = critical + high + normal
    return final_route, len(critical) + len(high)


def _route_total_distance(route, distances):
    if not route:
        return 0.0
    total_km = distances[0][route[0]]
    for idx in range(len(route) - 1):
        total_km += distances[route[idx]][route[idx + 1]]
    total_km += distances[route[-1]][0]
    return float(total_km)


def simulate_route(route_indices, locations, durations, driver, distances=None):
    results = []
    current_min = time_to_min(driver["shift_start"])
    max_end_min = current_min + driver["max_hours"] * 60
    prev_idx = 0

    for idx in route_indices:
        loc = locations[idx]
        travel_min = durations[prev_idx][idx]
        arrival_min = current_min + travel_min
        tw_start = time_to_min(loc["time_window"][0])
        tw_end = time_to_min(loc["time_window"][1])

        if arrival_min < tw_start:
            arrival_min = tw_start

        departure_min = arrival_min + SERVICE_TIME_MIN
        dist_from_prev = distances[prev_idx][idx] if distances else 0.0

        results.append({
            "id": loc["id"],
            "name": loc["name"],
            "priority": loc["priority"],
            "arrival": min_to_time(arrival_min),
            "departure": min_to_time(departure_min),
            "window": f"{loc['time_window'][0]}–{loc['time_window'][1]}",
            "is_late": arrival_min > tw_end,
            "late_by_min": max(0, int(arrival_min - tw_end)),
            "shift_exceeded": departure_min > max_end_min,
            "dist_from_prev": round(dist_from_prev, 2),
        })

        current_min = departure_min
        prev_idx = idx

    total_duration = current_min - time_to_min(driver["shift_start"])
    return results, total_duration


# ── Printing ─────────────────────────────────────────────────────────────────
def print_routes(vehicle_routes):
    print("\n" + "=" * 70)
    print("  FINAL DELIVERY PLAN")
    print("=" * 70)

    total_late = 0
    for entry in vehicle_routes:
        v = entry["vehicle"]
        driver = entry["driver"]
        stops = entry["stops"]
        km = entry["total_km"]
        dur = entry["total_duration_min"]

        if not stops:
            continue

        dname = driver["name"] if driver else "⚠ No driver"
        print(f"\n  🚐 {v['label']}  |  Driver: {dname}  |  {len(stops)} stops  |  {km:.1f} km  |  ~{int(dur)} min")
        print(f"  {'#':<4} {'Client':<28} {'Priority':<10} {'Window':<14} {'Arrival':<10} {'Status'}")
        print("  " + "-" * 72)

        for i, stop in enumerate(stops):
            priority_icon = "🔴" if stop["priority"] == "critical" else ("🟡" if stop["priority"] == "high" else "⚪")
            status = "🔴 LATE" if stop["is_late"] else "✅ OK"
            shift_warn = " ⚠ SHIFT EXCEEDED" if stop["shift_exceeded"] else ""
            late_note = f" (+{stop['late_by_min']} min)" if stop["is_late"] else ""
            dist_str = f"{stop['dist_from_prev']:.1f} km" if i > 0 else "start"
            print(f"  {i+1:<4} {stop['name']:<28} {priority_icon} {stop['priority']:<8} {stop['window']:<14} {stop['arrival']:<10} {dist_str:>8}  {status}{late_note}{shift_warn}")
            if stop["is_late"]:
                total_late += 1

        print(f"  {'':<4} {'↩ Return to Warehouse':<28}")

    print("\n" + "=" * 70)
    print("  SUMMARY")
    print("=" * 70)
    total_stops = sum(len(e["stops"]) for e in vehicle_routes)
    total_km = sum(e["total_km"] for e in vehicle_routes)
    print(f"  Total deliveries : {total_stops}")
    print(f"  Total distance   : {total_km:.1f} km")
    print(f"  Late deliveries  : {total_late}")
    if total_late == 0:
        print("  ✅ All deliveries within time windows")
    else:
        print(f"  ⚠  {total_late} deliveries missed their time window")
    print("=" * 70)


def print_route_comparison(vehicle_routes):
    print("\n" + "=" * 70)
    print("  ROUTING COMPARISON: RAW ORDER VS BUFFER-AWARE FINAL ROUTE")
    print("=" * 70)

    for entry in vehicle_routes:
        v = entry["vehicle"]
        driver = entry["driver"]
        dname = driver["name"] if driver else "⚠ No driver"
        raw_names = [entry["locations"][i]["name"] for i in entry["raw_route"]]
        final_names = [entry["locations"][i]["name"] for i in entry["optimized_route"]]
        raw_late = sum(1 for stop in entry["raw_stops"] if stop["is_late"])
        final_late = sum(1 for stop in entry["optimized_stops"] if stop["is_late"])

        print(f"\n  🚐 {v['label']}  |  Driver: {dname}")
        print(f"  Raw order    : {raw_names}")
        print(f"    Distance: {entry['raw_total_km']:.1f} km | Duration: ~{int(entry['raw_total_duration_min'])} min | Late: {raw_late}")
        print(f"  Final route  : {final_names}")
        print(f"    Distance: {entry['total_km']:.1f} km | Duration: ~{int(entry['total_duration_min'])} min | Late: {final_late}")

    print("=" * 70)


if __name__ == "__main__":
    locations = geocode_all_cached(ORDERS, WAREHOUSE)
    distances, durations = get_distance_matrix(locations)
    if distances is None:
        raise SystemExit(1)

    vehicle_orders = assign_orders_to_vehicles(ORDERS)
    driver_assignment = assign_drivers(vehicle_orders)

    print("\n" + "=" * 65)
    print("  STEP 4 — ROUTING (Critical first, then nearest-neighbor)")
    print("=" * 65)

    loc_index = {loc["id"]: i for i, loc in enumerate(locations)}
    vehicle_routes = []

    for v in VEHICLES:
        orders = vehicle_orders.get(v["id"], [])
        if not orders:
            continue

        driver = driver_assignment.get(v["id"])
        dname = driver["name"] if driver else "No driver"
        stop_indices = [loc_index[o["id"]] for o in orders if o["id"] in loc_index]

        print(f"\n  {v['label']} ({dname}) — {len(stop_indices)} stops")

        raw_route, raw_locked_count = build_route_without_optimization(stop_indices, locations)
        final_route, locked_count = build_route(stop_indices, locations, distances, driver, durations)
        print(f"    Raw order (no optimization): {[locations[i]['name'] for i in raw_route]}")
        print(f"    Locked (critical): {[locations[i]['name'] for i in final_route[:locked_count]]}")
        print(f"    Final route: {[locations[i]['name'] for i in final_route]}")

        raw_total_km = _route_total_distance(raw_route, distances)
        total_km = _route_total_distance(final_route, distances)
        raw_stops, raw_total_min = simulate_route(raw_route, locations, durations, driver or DRIVERS_DEFAULT[-1], distances)
        sim_stops, total_min = simulate_route(final_route, locations, durations, driver or DRIVERS_DEFAULT[-1], distances)

        vehicle_routes.append({
            "vehicle": v,
            "driver": driver,
            "locations": locations,
            "raw_route": raw_route,
            "raw_stops": raw_stops,
            "raw_total_km": raw_total_km,
            "raw_total_duration_min": raw_total_min,
            "optimized_route": final_route,
            "optimized_stops": sim_stops,
            "stops": sim_stops,
            "total_km": total_km,
            "total_duration_min": total_min,
        })

    print_route_comparison(vehicle_routes)
    print_routes(vehicle_routes)
